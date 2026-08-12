#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导出整个 ledger.db 为多 Sheet Excel"""
import sqlite3
import sys

sys.path.insert(0, "/opt/data/workspace/projects/parts-ledger")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = "/opt/data/workspace/projects/parts-ledger/ledger.db"
OUT = "/opt/data/workspace/projects/parts-ledger/进销存全库_2026-08-09.xlsx"

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for r in rows:
        ws.append(r)
    for c in range(1, len(headers) + 1):
        width = max(
            len(str(headers[c - 1])),
            max((len(str(r[c - 1])) for r in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(c)].width = min(width + 4, 40)


# Sheet1 流水
ws = wb.active
ws.title = "流水"
rows = conn.execute(
    """SELECT t.biz_date, t.biz_type, p.name, p.model, p.unit, t.qty,
              t.price, t.amount, c.name AS cp_name, c.region, c.phone, t.note
       FROM transactions t
       JOIN products p ON t.product_id = p.id
       LEFT JOIN counterparties c ON t.counterparty_id = c.id
       ORDER BY t.biz_date, t.id"""
).fetchall()
data = []
for r in rows:
    kind = "进货" if r["biz_type"] == "purchase" else "售出"
    data.append(
        [r["biz_date"], kind, r["name"], r["model"], r["unit"], r["qty"],
         r["price"], r["amount"], r["cp_name"] or "", r["region"] or "",
         r["phone"] or "", r["note"] or ""]
    )
write_sheet(ws, ["日期", "类型", "名称", "型号", "单位", "数量", "单价", "金额",
                 "往来单位", "地区", "电话", "备注"], data)
# 售出合计
sales = [r for r in rows if r["biz_type"] == "sale"]
ws.append(["", "售出合计", "", "", "", "", "",
           round(sum(r["amount"] for r in sales), 2), "", "", "", ""])
ws.append(["", "进货合计", "", "", "", "", "",
           round(sum(r["amount"] for r in rows if r["biz_type"] == "purchase"), 2),
           "", "", "", ""])

# Sheet2 商品
ws2 = wb.create_sheet("商品")
rows = conn.execute("SELECT * FROM products ORDER BY name, model").fetchall()
data = [[r["id"], r["name"], r["model"], r["unit"], r["cost_price"] or "",
         r["sale_price"] or ""] for r in rows]
write_sheet(ws2, ["ID", "名称", "型号", "单位", "参考进价", "参考售价"], data)

# Sheet3 客户
ws3 = wb.create_sheet("客户")
rows = conn.execute("SELECT * FROM counterparties ORDER BY ctype, name").fetchall()
data = [[r["id"], r["name"], r["region"] or "", r["phone"] or "",
         "供应商" if r["ctype"] == "supplier" else "客户"] for r in rows]
write_sheet(ws3, ["ID", "名称", "地区", "电话", "类型"], data)

# Sheet4 库存
ws4 = wb.create_sheet("库存")
rows = conn.execute(
    """SELECT p.name, p.model, p.unit,
              SUM(CASE WHEN t.biz_type='purchase' THEN t.qty
                       WHEN t.biz_type='sale' THEN -t.qty ELSE 0 END) AS stock,
              p.cost_price
       FROM products p LEFT JOIN transactions t ON t.product_id = p.id
       GROUP BY p.id ORDER BY p.name, p.model"""
).fetchall()
data = [[r["name"], r["model"], r["unit"], r["stock"] or 0,
         r["cost_price"] or ""] for r in rows]
write_sheet(ws4, ["名称", "型号", "单位", "当前库存", "参考进价"], data)

wb.save(OUT)
print(f"已导出: {OUT}")
print(f"流水 {len([r for r in conn.execute('SELECT 1 FROM transactions')])} 条, "
      f"商品 {len([r for r in conn.execute('SELECT 1 FROM products')])} 条, "
      f"客户 {len([r for r in conn.execute('SELECT 1 FROM counterparties')])} 条")
