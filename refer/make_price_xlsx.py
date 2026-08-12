#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把配件价格表做成 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    (1, "螺母", "200-01-013", 85.00),
    (2, "防护罩", "L250-01-066", 12.00),
    (3, "O型橡胶圈", "200-01-039", 70.00),
    (4, "旋阀器组件", "200-03-500", 270.00),
    (5, "进气阀座", "200-03-022", 150.00),
    (6, "排气阀座", "200-03-024", 215.00),
    (7, "气阀导管", "200-03-008", 70.00),
    (8, "喷油器压板螺栓", "200-03-012", 19.00),
    (9, "高压油管组件", "200-53-000", 670.00),
    (10, "滑油双联滤清器罩壳底部密封圈", "", 10.00),
    (11, "气缸盖罩密封圈", "200-03-401", 70.00),
    (12, "气阀锁片", "200-03-015", 13.00),
    (13, "派克精密O型圈", "200-03-026", 13.00),
    (14, "派克精密O型圈", "200-03-027", 12.00),
]

wb = Workbook()
ws = wb.active
ws.title = "配件价格表"

headers = ["序号", "名称", "图号", "单价", "合计"]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
body_font = Font(name="微软雅黑", size=11)
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

# 表头
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border

# 数据
for i, (no, name, partno, price) in enumerate(rows, start=2):
    vals = [no, name, partno if partno else "[无]", price, None]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = body_font
        c.border = border
        c.alignment = center if col != 2 else left
    if i % 2 == 0:
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = PatternFill(
                start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
            )

# 合计行（单价总和）
total = sum(r[3] for r in rows)
r = len(rows) + 2
ws.cell(row=r, column=1, value="合计")
ws.cell(row=r, column=4, value=round(total, 2))
for col in range(1, 6):
    c = ws.cell(row=r, column=col)
    c.font = Font(name="微软雅黑", size=11, bold=True)
    c.border = border
    c.alignment = center
    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# 列宽
widths = [8, 34, 16, 12, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结首行
ws.freeze_panes = "A2"

out = "/opt/data/workspace/outputs/柴油机配件价格表_2026-08-10.xlsx"
wb.save(out)
print("saved:", out)
print("单价合计:", total)
