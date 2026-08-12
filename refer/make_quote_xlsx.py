#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""报价单图片 → Excel：合并第10/11行为一行（滑油双联滤器罩壳底部密封圈）"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

headers = ["序号", "名称", "图号", "单价", "合计"]
rows = [
    (1,  "螺母",                      "200-01-013", 85.00, None),
    (2,  "防护罩",                    "L250-01-066", 12.00, None),
    (3,  "O型橡胶圈",                 "200-01-039", 70.00, None),
    (4,  "旋阀器组件",                "200-03-500", 270.00, None),
    (5,  "进气阀座",                  "200-03-022", 150.00, None),
    (6,  "排气阀座",                  "200-03-024", 215.00, None),
    (7,  "气阀导管",                  "200-03-008", 70.00, None),
    (8,  "喷油器压板螺栓",            "200-03-012", 19.00, None),
    (9,  "高压油管组件",              "200-53-000", 670.00, None),
    (10, "滑油双联滤器罩壳底部密封圈", None,         10.00, None),
    (11, "气缸盖罩密封圈",            "200-03-401", 70.00, None),
    (12, "气阀锁片",                  "200-03-015", 13.00, None),
    (13, "派克精密O型圈",             "200-03-026", 13.00, None),
    (14, "派克精密O型圈",             "200-03-027", 12.00, None),
]

wb = Workbook()
ws = wb.active
ws.title = "报价明细"

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="D9E1F2")

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = header_fill
    c.border = border

for i, row in enumerate(rows, start=2):
    for col, v in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.border = border
        if isinstance(v, (int, float)):
            c.alignment = Alignment(horizontal="center")
            if isinstance(v, float):
                c.number_format = "0.00"

widths = [6, 26, 14, 12, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out = "/opt/data/workspace/outputs/报价明细.xlsx"
wb.save(out)
print("saved:", out)

# 验证读回
wb2 = load_workbook(out)
ws2 = wb2.active
print("验证 行数:", ws2.max_row, "列数:", ws2.max_column)
for r in range(1, ws2.max_row + 1):
    print("  行", r, [ws2.cell(row=r, column=c).value for c in range(1, 6)])
