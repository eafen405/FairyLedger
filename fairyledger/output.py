"""文件产出：Excel 导出/周期汇总写入与库文件备份（spec D6「导出与备份」）。

- openpyxl 延迟导入：未安装时仅 export/report period 触发系统错误（ImportError →
  CLI 退出码 2），其余命令不受影响。
- 导出 = 整库快照单工作簿 5 Sheet（流水/商品/往来单位/库存/毛利，spec D6）；
  周期汇总 = 期间汇总/按产品分组/按客户分组/流水明细 4 Sheet（spec D6 周期汇总四块）。
- 备份 = 复制库文件到库同目录 backups/，保留最近 7 份轮转（ledger-YYYYMMDD.db，
  同日同名覆盖天然幂等，超出删除最旧）。
- 文件产出命令天然可重跑（幂等），无事务原子性要求。
- 查询类 --out（spec D5）：查询结果同时落 xlsx（list 单 sheet「结果」、dict 每个
  list 值一个同名 sheet），CLI 只作副作用不 emit 文件信息。
"""

import datetime
import shutil
from pathlib import Path
from typing import Any, Sequence

# 类型中文标签（导出/周期汇总明细共用；日报/查询输出仍用英文 biz_type）
BIZ_TYPE_LABEL = {
    "purchase": "进货",
    "sale": "售出",
    "purchase_return": "进货退货",
    "sale_return": "售出退货",
}


def default_export_path() -> Path:
    """默认导出文件名 FairyLedger_YYYYMMDD.xlsx（导出时点当天）。"""
    return Path.cwd() / f"FairyLedger_{datetime.date.today().strftime('%Y%m%d')}.xlsx"


def default_period_path(from_date: str, to_date: str) -> Path:
    """周期汇总默认文件名带区间：FairyLedger_周期汇总_YYYYMMDD-YYYYMMDD.xlsx。"""
    return Path.cwd() / (
        f"FairyLedger_周期汇总_{from_date.replace('-', '')}-{to_date.replace('-', '')}.xlsx"
    )


def _workbook():
    """延迟导入 openpyxl；缺失时让 ImportError 冒泡 → CLI 归为系统错误退出码 2。"""
    from openpyxl import Workbook

    return Workbook()


def _write_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    """表头 + 数据行；None 写为空串（避免 Excel 显示 'None'）。"""
    ws.append(list(headers))
    for row in rows:
        ws.append(["" if v is None else v for v in row])


def _autosize(ws, rows: Sequence[Sequence[Any]], ncols: int) -> None:
    """按表头与数据行内容估算列宽（min 宽度 4、max 40，参考 refer/ 历史脚本）。"""
    from openpyxl.utils import get_column_letter

    for c in range(1, ncols + 1):
        width = max(
            [len(str(ws.cell(row=1, column=c).value or ""))]
            + [len(str(r[c - 1])) for r in rows if c - 1 < len(r)]
        )
        ws.column_dimensions[get_column_letter(c)].width = min(width + 4, 40)


def _qty_unit(qty: float, unit: str) -> str:
    """数量+单位：整数值去掉浮点尾巴（5.0 → '5 件'），小数保留（2.5 → '2.5 件'）。"""
    return f"{int(qty)} {unit}" if qty == int(qty) else f"{qty} {unit}"


def _ref_marker(ref: dict) -> str:
    """红冲原单标记：'#<原单 id> <原单日期> <原单类型中文>'。"""
    return f"#{ref['id']} {ref['biz_date']} {BIZ_TYPE_LABEL[ref['biz_type']]}"


def _flow_row(r: dict) -> list:
    """导出「流水」sheet 行：日期/类型/图号/名称/数量+单位/单价/金额/运费/成本/
    往来单位/红冲原单/备注（spec D6 导出流水列）。"""
    return [
        r["biz_date"], BIZ_TYPE_LABEL[r["biz_type"]], r["drawing_no"], r["name"],
        _qty_unit(r["qty"], r["unit"]), r["price"], r["amount"], r["freight"],
        r["cost"], r["counterparty"] or "",
        _ref_marker(r["ref"]) if r["ref"] else "", r["note"] or "",
    ]


def _detail_row(r: dict) -> list:
    """周期汇总「流水明细」sheet 行（同日报明细结构）：日期/类型/图号/名称/
    数量+单位/单价/金额/往来单位 + 退货行「红冲」原单标记。"""
    return [
        r["biz_date"], BIZ_TYPE_LABEL[r["biz_type"]], r["drawing_no"], r["name"],
        _qty_unit(r["qty"], r["unit"]), r["price"], r["amount"],
        r["counterparty"] or "",
        _ref_marker(r["ref"]) if r["ref"] else "",
    ]


def _write_workbook(
    sheets: Sequence[tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
    out_path: Path,
) -> dict:
    """单工作簿写入（export/period 共用）：逐 sheet 写表头与数据、估算列宽，
    保存后返回 {file, sheets:[{name, rows}]}；None 值写为空单元格。"""
    wb = _workbook()
    for i, (name, headers, rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name
        _write_sheet(ws, headers, rows)
        _autosize(ws, rows, len(headers))
    out_path = Path(out_path).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return {
        "file": str(out_path),
        "sheets": [{"name": name, "rows": len(rows)} for name, _, rows in sheets],
    }


def write_export_xlsx(data: dict, out_path: Path) -> dict:
    """整库快照写入 Excel 单工作簿 5-Sheet（spec D6 导出列定义）。"""
    sheets = [
        ("流水", ["日期", "类型", "图号", "名称", "数量+单位", "单价", "金额",
                 "运费", "成本", "往来单位", "红冲原单", "备注"],
         [_flow_row(r) for r in data["transactions"]]),
        ("商品", ["图号", "名称", "单位", "别名"],
         [[r["drawing_no"], r["name"], r["unit"], r["aliases"]]
          for r in data["products"]]),
        ("往来单位", ["名称", "类型标记", "挂账标记", "联系方式"],
         [[r["name"], r["roles"], "是" if r["is_credit"] else "否", r["contact"]]
          for r in data["counterparties"]]),
        ("库存", ["图号", "名称", "数量", "当前均价", "金额"],
         [[r["drawing_no"], r["name"], r["qty"], r["unit_cost"], r["amount"]]
          for r in data["stock"]]),
        ("毛利", ["图号", "名称", "售出金额", "成本", "毛利", "毛利率"],
         [[r["drawing_no"], r["name"], r["amount"], r["cost"], r["margin"],
           r["margin_rate"]]
          for r in data["margin"]]),
    ]
    return _write_workbook(sheets, out_path)


def write_period_xlsx(data: dict, out_path: Path) -> dict:
    """周期汇总写入 Excel：期间汇总/按产品分组/按客户分组/流水明细 4-Sheet。

    TOP 10 仅为参考（spec D6）：按产品/按客户分组取毛利降序前 10。
    """
    def _row(label: str, value: Any) -> list:
        return [label, value]

    summary_rows = [
        _row("进货笔数", data["summary"]["purchase"]["count"]),
        _row("进货金额", data["summary"]["purchase"]["amount"]),
        _row("进货运费", data["summary"]["purchase"]["freight"]),
        _row("售出笔数", data["summary"]["sale"]["count"]),
        _row("售出金额", data["summary"]["sale"]["amount"]),
        _row("售出运费", data["summary"]["sale"]["freight"]),
        _row("毛利", data["summary"]["margin"]),
        _row("期初库存金额", data["summary"]["opening_inventory"]),
        _row("期末库存金额", data["summary"]["closing_inventory"]),
    ]
    top_products = [
        [r["drawing_no"], r["name"], r["sale_qty"], r["amount"], r["cost"],
         r["margin"], r["margin_rate"]]
        for r in data["by_product"][:10]
    ]
    top_customers = [
        [r["customer"] or "", r["amount"], r["margin"], r["margin_rate"]]
        for r in data["by_customer"][:10]
    ]
    sheets = [
        ("期间汇总", ["项目", "金额"], summary_rows),
        ("按产品分组", ["图号", "名称", "售出数量", "售出金额", "成本", "毛利", "毛利率"],
         top_products),
        ("按客户分组", ["客户", "售出金额", "毛利", "毛利率"], top_customers),
        ("流水明细", ["日期", "类型", "图号", "名称", "数量+单位", "单价", "金额",
                    "往来单位", "红冲原单"],
         [_detail_row(r) for r in data["details"]]),
    ]
    return _write_workbook(sheets, out_path)


def _query_cell(v: Any) -> Any:
    """查询落表单元格值：list/tuple（如检索结果的别名数组）拼 '、' 串，其余原样
    （None 由 _write_sheet 置空串；与 export 商品 sheet 别名列口径一致）。"""
    if isinstance(v, (list, tuple)):
        return "、".join(str(x) for x in v)
    return v


def write_query_xlsx(payload: Any, out_path: Path) -> dict:
    """查询结果落 Excel（spec D5：查询类 --out 可选文件产出，side-effect 不 emit）。

    list（query stock/price/history）→ 单 sheet「结果」：表头 = 首元素 keys
    （插入序），每行 = 各 dict 的 values 按同序；空列表无表头只留空 sheet。
    dict（query credit/product/margin）→ 每个「值为 list」的 key 一个同名 sheet
    （records/balances/products/by_product/by_customer），跳过标量 key
    （match/from/to/amount/cost/margin/margin_rate）；match:none 时 products 为
    空列表仍写空 sheet。list 型单元格值（如检索结果 aliases 数组）拼 '、' 串。
    返回值形状对齐 _write_workbook，调用方 CLI 不 emit。
    """
    if isinstance(payload, dict):
        sheets = []
        for name, rows in payload.items():
            if not isinstance(rows, list):
                continue
            headers = list(rows[0].keys()) if rows else []
            data = [[_query_cell(r[k]) for k in headers] for r in rows]
            sheets.append((name, headers, data))
    else:
        rows = list(payload) if payload else []
        headers = list(rows[0].keys()) if rows else []
        data = [[_query_cell(r[k]) for k in headers] for r in rows]
        sheets = [("结果", headers, data)]
    return _write_workbook(sheets, out_path)


def backup_db(db_path: Path) -> dict:
    """备份库文件：复制到库同目录 backups/ledger-YYYYMMDD.db，保留最近 7 份轮转。

    同日多次备份同名覆盖（天然幂等）；超出 7 份按文件名日期序删除最旧。
    """
    db_path = Path(db_path)
    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.date.today().strftime("%Y%m%d")
    target = backup_dir / f"ledger-{today}.db"
    shutil.copy2(db_path, target)
    backups = sorted(backup_dir.glob("ledger-*.db"), key=lambda p: p.name)
    removed: list[str] = []
    while len(backups) > 7:
        oldest = backups.pop(0)
        oldest.unlink()
        removed.append(oldest.name)
    return {
        "backup": str(target),
        "retained": [p.name for p in backups],
        "removed": removed,
    }
